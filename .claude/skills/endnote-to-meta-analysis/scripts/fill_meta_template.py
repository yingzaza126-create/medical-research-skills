#!/usr/bin/env python3
"""
fill_meta_template.py -- write extracted study data into
`1.1Data_templet_for_Meta.xlsx` WITHOUT destroying its structure.

It opens the ORIGINAL template (keeping sheet name, header row, and all
formatting), then appends one row per extracted record starting at row 2.
It never renames or deletes template columns; it only maps incoming fields
onto the columns that already exist in the header row.

Input data = a CSV (or JSON list) whose column names match the template
headers (Study_ID, Author, Year, ... ). Any header not present in the input
is left blank. Missing values are written verbatim -- use NR / NA / NE /
"Not poolable" in your source data; this script never fabricates numbers.

Usage:
    python fill_meta_template.py \
        --template assets/templates/1.1Data_templet_for_Meta.xlsx \
        --data extracted_meta.csv \
        --out  1.1Data_templet_for_Meta_FILLED.xlsx

Validation performed (reports only; nothing is auto-corrected):
    - duplicate Study_ID
    - events > sample size  (Success_* vs Sample_Size_*)
    - CI_lower > OR > CI_upper ordering
    - negative / zero SD
    - Excel error strings (#REF!, #DIV/0!, ...) left in template
"""
import argparse
import csv
import json
import sys

import openpyxl

ERR = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


def load_data(path):
    if path.lower().endswith(".json"):
        return json.load(open(path, encoding="utf-8"))
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def to_num(v):
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--template", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--start-row", type=int, default=2)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.template)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None:
            headers[str(h).strip()] = c
    print(f"Template sheet '{ws.title}' with {len(headers)} columns:")
    print("  " + ", ".join(headers))

    rows = load_data(a.data)
    unknown = set()
    for r in rows:
        for k in r:
            if k not in headers:
                unknown.add(k)
    if unknown:
        print(f"\nWARNING: {len(unknown)} input field(s) not in template header "
              f"(left out, template not modified): {sorted(unknown)}")

    r_out = a.start_row
    for rec in rows:
        for h, col in headers.items():
            if h in rec and rec[h] not in (None, ""):
                ws.cell(r_out, col).value = rec[h]
        r_out += 1

    # ---- validation (report only) ----
    flags = []
    seen = {}
    for r in range(a.start_row, r_out):
        sid = ws.cell(r, headers.get("Study_ID", 1)).value
        if sid is not None:
            if sid in seen:
                flags.append(f"row {r}: duplicate Study_ID '{sid}' (also row {seen[sid]})")
            seen[sid] = r

        def cell(name):
            return ws.cell(r, headers[name]).value if name in headers else None

        for ev, ss in (("Success_Intervention", "Sample_Size_Intervention"),
                       ("Success_Control", "Sample_Size_Control")):
            e, n = to_num(cell(ev)), to_num(cell(ss))
            if e is not None and n is not None and e > n:
                flags.append(f"row {r}: {ev}={e} > {ss}={n} (events exceed sample)")
        orr, lo, hi = to_num(cell("OR")), to_num(cell("CI_95_Lower")), to_num(cell("CI_95_Upper"))
        if None not in (orr, lo, hi) and not (lo <= orr <= hi):
            flags.append(f"row {r}: CI ordering off (lower={lo}, OR={orr}, upper={hi})")
        for sd in ("SD_Intervention", "SD_Control"):
            s = to_num(cell(sd))
            if s is not None and s <= 0:
                flags.append(f"row {r}: {sd}={s} not > 0")
        for h, col in headers.items():
            v = ws.cell(r, col).value
            if isinstance(v, str) and v in ERR:
                flags.append(f"row {r}: Excel error '{v}' in column {h}")

    wb.save(a.out)
    print(f"\nWrote {r_out - a.start_row} data row(s) -> {a.out}")
    if flags:
        print(f"\nVALIDATION FLAGS ({len(flags)}) -- require human verification:")
        for fl in flags:
            print("  FLAG:", fl)
    else:
        print("\nValidation: no numerical inconsistencies detected.")


if __name__ == "__main__":
    main()
