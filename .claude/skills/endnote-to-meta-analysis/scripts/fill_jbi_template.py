#!/usr/bin/env python3
"""
fill_jbi_template.py -- write JBI critical-appraisal results into
`1S1_Table_JBI_Quality_Appraisal.xlsx`, preserving the template's title
rows and two-row header (rows 1-4). Data is written from row 5 down.

The template's fixed layout (do not change):
    A Study (#)      B Author, Year   C Study Design   D JBI Tool Used
    E-J Core appraisal domains (row-4 sub-labels):
        E Eligibility Criteria         F Participants Representative
        G Exposure Measured Validly    H Confounders Identified
        I Outcomes Measured Validly    J Statistical Analysis Appropriate
    K Additional Domains Met   L Total Items Applicable   M Items Met (Yes)
    N JBI Score (%)            O Quality Rating

Input CSV/JSON keys (any missing key -> left blank):
    Study, Author_Year, Study_Design, JBI_Tool,
    Domain_1..Domain_6            (Y / N / Unclear / NA -- per checklist item)
    Additional_Domains_Met, Total_Items_Applicable, Items_Met,
    JBI_Score, Quality_Rating

Rules honoured:
    * JBI_Score is arithmetic: if left blank AND Items_Met + Total_Items_
      Applicable are numeric, it is computed = round(100*Items/Total).
      Nothing else is derived.
    * Quality_Rating is NEVER auto-assigned (no official universal % cut-off);
      it is written only if supplied. Use --rating-cutoffs to opt in.
    * A domain must be graded from evidence in the source; this script only
      transcribes what you provide. Blank domains stay blank (=> Unclear/NR
      is your call, recorded in your input).

Usage:
    python fill_jbi_template.py \
        --template assets/templates/1S1_Table_JBI_Quality_Appraisal.xlsx \
        --data jbi_appraisal.csv \
        --out  1S1_Table_JBI_Quality_Appraisal_FILLED.xlsx
"""
import argparse
import csv
import json

import openpyxl
from openpyxl.cell.cell import MergedCell

COLS = {
    "Study": 1, "Author_Year": 2, "Study_Design": 3, "JBI_Tool": 4,
    "Domain_1": 5, "Domain_2": 6, "Domain_3": 7, "Domain_4": 8,
    "Domain_5": 9, "Domain_6": 10, "Additional_Domains_Met": 11,
    "Total_Items_Applicable": 12, "Items_Met": 13, "JBI_Score": 14,
    "Quality_Rating": 15,
}


def load_data(path):
    if path.lower().endswith(".json"):
        return json.load(open(path, encoding="utf-8"))
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def num(v):
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--template", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--start-row", type=int, default=5)
    ap.add_argument("--rating-cutoffs", default="",
                    help="optional 'high,moderate' %% cut-offs, e.g. '80,50' "
                         "to derive Quality_Rating from JBI_Score when blank")
    a = ap.parse_args()

    cut = None
    if a.rating_cutoffs:
        hi, mod = (float(x) for x in a.rating_cutoffs.split(","))
        cut = (hi, mod)

    from copy import copy

    wb = openpyxl.load_workbook(a.template)
    ws = wb.active
    rows = load_data(a.data)
    n = len(rows)
    flags = []

    def put(r, c, v):
        cell = ws.cell(r, c)
        if isinstance(cell, MergedCell):
            return False
        cell.value = v
        return True

    # ---- locate a SUMMARY/footer block (merged rows at/after start_row) ----
    footer_merges = [(cr.min_row, cr.min_col, cr.max_row, cr.max_col)
                     for cr in ws.merged_cells.ranges if cr.min_row >= a.start_row]
    first_footer = min((r for r, *_ in footer_merges), default=None)
    footer_style = None
    if first_footer is not None:
        # remember styling + column spans of the footer, then remove it
        footer_style = {c: (copy(ws.cell(first_footer, c).font),
                            copy(ws.cell(first_footer, c).fill),
                            copy(ws.cell(first_footer, c).alignment),
                            ws.cell(first_footer, c).number_format)
                        for c in range(1, 16)
                        if not isinstance(ws.cell(first_footer, c), MergedCell)}
        spans = [(c1, c2) for (r1, c1, r2, c2) in footer_merges if r1 == first_footer]
        for (r1, c1, r2, c2) in footer_merges:
            ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        for r in range(first_footer, first_footer + 1):
            for c in range(1, 16):
                put(r, c, None)

    # clear existing example/placeholder value cells (keep formatting)
    for r in range(a.start_row, max(ws.max_row, a.start_row + n) + 1):
        for c in range(1, 16):
            put(r, c, None)

    r_out = a.start_row
    for i, rec in enumerate(rows, 1):
        if "Study" not in rec or rec.get("Study") in (None, ""):
            rec["Study"] = i
        # arithmetic score only
        if not rec.get("JBI_Score"):
            im, ta = num(rec.get("Items_Met")), num(rec.get("Total_Items_Applicable"))
            if im is not None and ta and ta > 0:
                if im > ta:
                    flags.append(f"row {r_out}: Items_Met {im} > Total {ta}")
                rec["JBI_Score"] = round(100 * im / ta)
        if cut and not rec.get("Quality_Rating") and num(rec.get("JBI_Score")) is not None:
            s = num(rec["JBI_Score"])
            rec["Quality_Rating"] = ("High" if s >= cut[0]
                                     else "Moderate" if s >= cut[1] else "Low")
        for key, col in COLS.items():
            if key in rec and rec[key] not in (None, ""):
                put(r_out, col, rec[key])
        r_out += 1

    # ---- rebuild the SUMMARY footer one blank row below the data ----
    if first_footer is not None:
        fr = r_out + 1
        ratings = [str(rec.get("Quality_Rating") or "").strip().lower() for rec in rows]
        got = sum(1 for x in ratings if x in ("high", "moderate", "low"))
        if got == n and n:
            hi = ratings.count("high"); mod = ratings.count("moderate"); lo = ratings.count("low")
            summary = (f"High quality: {hi}/{n} ({round(100*hi/n)}%)   |   "
                       f"Moderate quality: {mod}/{n} ({round(100*mod/n)}%)   |   "
                       f"Low quality: {lo}/{n} ({round(100*lo/n)}%)")
        else:
            summary = ("Quality ratings pending full-text appraisal "
                       "(High/Moderate/Low = NR)")
        put(fr, 1, f"SUMMARY (n = {n} studies)")
        put(fr, 5, summary)
        if footer_style:
            for c, (font, fill, align, numfmt) in footer_style.items():
                cell = ws.cell(fr, c)
                cell.font = copy(font); cell.fill = copy(fill)
                cell.alignment = copy(align); cell.number_format = numfmt
        for (c1, c2) in spans:
            ws.merge_cells(start_row=fr, start_column=c1, end_row=fr, end_column=c2)

    wb.save(a.out)
    print(f"Wrote {r_out - a.start_row} appraisal row(s) -> {a.out}")
    if not cut:
        print("Quality_Rating written only where supplied "
              "(no cut-offs given; pass --rating-cutoffs to derive it).")
    if flags:
        print(f"\nVALIDATION FLAGS ({len(flags)}):")
        for fl in flags:
            print("  FLAG:", fl)


if __name__ == "__main__":
    main()
